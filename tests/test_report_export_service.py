from datetime import UTC, datetime
import json
from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from app.cli.main import app
from app.services.report_export_service import ReportExportService, build_summary_rows
from tests.test_reconciliation_service import make_folder, make_tir

runner = CliRunner()


def write_inputs(tmp_path: Path) -> tuple[Path, Path, Path]:
    tir_records = [
        make_tir(registry_file_ref="MISE-ABED-001", project_name="Community hall roof repair"),
        make_tir(
            registry_file_ref="MISE-CED-002",
            project_name="Road drainage",
            contact_email="road@example.test",
        ),
    ]
    folders = [
        make_folder(),
        make_folder(
            registry_file_ref="MISE-CED-002",
            project_name="Road drainage",
            path="C:/MISE/MISE-CED-002 - Road drainage",
        ),
    ]
    reconciliation = [
        {
            "registry_file_ref": "MISE-ABED-001",
            "project_name": "Community hall roof repair",
            "contact_email": "contact@example.test",
            "category": "MATCHED",
            "confidence_score": 100,
            "matched_folder_path": "C:/MISE/MISE-ABED-001 - Community hall roof repair",
            "reasons": ["Exact registry reference match"],
        },
        {
            "registry_file_ref": "MISE-CED-002",
            "project_name": "Road drainage",
            "contact_email": "road@example.test",
            "category": "MISSING_FOLDER",
            "confidence_score": 0,
            "matched_folder_path": None,
            "reasons": ["No folder matched registry reference or project name"],
        },
    ]
    tir_path = tmp_path / "tir_rows.json"
    folders_path = tmp_path / "folder_inventory.json"
    reconciliation_path = tmp_path / "reconciliation.json"
    tir_path.write_text(
        json.dumps({"normalized": [record.model_dump(mode="json") for record in tir_records]}),
        encoding="utf-8",
    )
    folders_path.write_text(
        json.dumps({"inventory": [folder.model_dump(mode="json") for folder in folders]}),
        encoding="utf-8",
    )
    reconciliation_path.write_text(json.dumps(reconciliation), encoding="utf-8")
    return tir_path, folders_path, reconciliation_path


def test_summary_rows_include_required_counts() -> None:
    summary_rows = build_summary_rows(
        [make_tir(), make_tir(registry_file_ref="MISE-CED-002", project_name="Road drainage")],
        [],
    )

    assert {"section": "project_status", "value": "IN_PROGRESS", "count": 2} in summary_rows
    assert {"section": "department_hod", "value": "ABED", "count": 2} in summary_rows


def test_report_export_generates_workbook_json_and_csv(tmp_path: Path) -> None:
    tir_path, folders_path, reconciliation_path = write_inputs(tmp_path)
    service = ReportExportService(timestamp=datetime(2026, 6, 9, tzinfo=UTC))

    result = service.export(
        tir_path=tir_path,
        folders_path=folders_path,
        reconciliation_path=reconciliation_path,
        out_dir=tmp_path,
    )

    assert Path(result.xlsx_path).name == "mise_report_20260609T000000Z.xlsx"
    assert Path(result.json_path).exists()
    assert Path(result.csv_path).exists()

    workbook = load_workbook(result.xlsx_path)
    assert workbook.sheetnames == [
        "Summary",
        "TIR Records",
        "Folder Inventory",
        "Reconciliation Results",
        "Data Quality Issues",
    ]

    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    data_quality_rows = list(workbook["Data Quality Issues"].iter_rows(values_only=True))
    assert ("project_status", "IN_PROGRESS", 2) in summary_rows
    assert ("reconciliation_category", "MATCHED", 1) in summary_rows
    assert ("reconciliation_category", "MISSING_FOLDER", 1) in summary_rows
    assert ("MISE-CED-002", "missing_funding_source", "WARNING", "TIR record is missing funding source") in data_quality_rows


def test_report_export_cli_outputs_paths(tmp_path: Path) -> None:
    tir_path, folders_path, reconciliation_path = write_inputs(tmp_path)
    out_dir = tmp_path / "reports"

    result = runner.invoke(
        app,
        [
            "report",
            "export",
            "--tir",
            str(tir_path),
            "--folders",
            str(folders_path),
            "--reconciliation",
            str(reconciliation_path),
            "--out",
            str(out_dir),
        ],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["summary_rows"] > 0
    assert Path(payload["xlsx_path"]).exists()
