from datetime import UTC, datetime
import json
from pathlib import Path

from openpyxl import load_workbook
from typer.testing import CliRunner

from app.cli.main import app
from app.domain.tir import TechnicalIntakeRequest
from app.services.data_quality_service import DataQualityService, export_data_quality_report

runner = CliRunner()


def make_record(**overrides: object) -> TechnicalIntakeRequest:
    values = {
        "created": datetime(2026, 6, 1, tzinfo=UTC),
        "secretary_approval": True,
        "mise_hod": "ABED",
        "registry_confirmation": True,
        "registry_file_ref": "MISE-ABED-001",
        "organisation": "Betio Town Council",
        "project_name": "Community hall roof repair",
        "service_request": "Inspection",
        "project_location": "Betio",
        "project_status": "IN_PROGRESS",
        "contact_person": "Example Contact",
        "contact_email": "contact@example.test",
        "funding_source": "Government",
    }
    values.update(overrides)
    return TechnicalIntakeRequest.model_construct(**values)


def issue_types(record: TechnicalIntakeRequest) -> set[str]:
    report = DataQualityService().check([record])
    return {issue.issue_type for issue in report.issues}


def test_data_quality_checks_every_issue_type() -> None:
    record = make_record(
        project_name="",
        contact_email="",
        registry_file_ref="",
        funding_source="",
        project_status="WAITING",
        service_request="",
        project_location="",
        secretary_approval=None,
        registry_confirmation=None,
    )

    assert issue_types(record) == {
        "missing_project_name",
        "missing_contact_email",
        "missing_registry_file_reference",
        "missing_funding_source",
        "invalid_project_status",
        "missing_service_request",
        "missing_project_location",
        "missing_secretary_approval_value",
        "missing_registry_confirmation_value",
    }


def test_data_quality_completeness_score_and_summary() -> None:
    report = DataQualityService().check([make_record(funding_source="")])

    assert report.records[0].completeness_score == 89
    assert report.summary.issue_counts == {"missing_funding_source": 1}
    assert report.summary.severity_counts == {"WARNING": 1}


def test_data_quality_export_xlsx(tmp_path: Path) -> None:
    output_path = tmp_path / "data_quality.xlsx"
    report = DataQualityService().check([make_record(funding_source="")])

    export_data_quality_report(report, output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Summary", "Issues", "Records"]
    issue_rows = list(workbook["Issues"].iter_rows(values_only=True))
    assert ("MISE-ABED-001", "missing_funding_source", "WARNING", "TIR record is missing funding source") in issue_rows


def test_data_quality_cli_outputs_summary(tmp_path: Path) -> None:
    tir_path = tmp_path / "tir_rows.json"
    out_path = tmp_path / "data_quality.xlsx"
    tir_path.write_text(
        json.dumps({"normalized": [make_record(funding_source="").model_dump(mode="json")]}),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        ["data-quality", "check", "--tir", str(tir_path), "--out", str(out_path)],
    )

    assert result.exit_code == 0
    payload = json.loads(result.stdout)
    assert payload["issue_counts"] == {"missing_funding_source": 1}
    assert out_path.exists()
